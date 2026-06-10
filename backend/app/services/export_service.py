import csv
import os
import tempfile
import zipfile
from io import BytesIO
from pathlib import Path

from app.core.config import settings


def _find_cjk_font() -> str:
    """查找系统中可用的中文字体文件，返回路径"""
    candidates = [
        # Linux (Docker)
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        # macOS
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        # Windows
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return ""


def export_excel(project, shots) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"

    headers = ["镜头号", "镜头类型", "时长(秒)", "画面描述", "氛围", "对应脚本", "AI提示词", "备注"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, shot in enumerate(shots, 2):
        ws.cell(row=row_idx, column=1, value=shot.shot_number)
        ws.cell(row=row_idx, column=2, value=shot.shot_type or "")
        ws.cell(row=row_idx, column=3, value=shot.duration_sec or "")
        ws.cell(row=row_idx, column=4, value=shot.content or "")
        ws.cell(row=row_idx, column=5, value=shot.atmosphere or "")
        ws.cell(row=row_idx, column=6, value=shot.script_reference or "")
        ws.cell(row=row_idx, column=7, value=shot.ai_prompt or "")
        ws.cell(row=row_idx, column=8, value=shot.notes or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 25

    filepath = os.path.join(settings.UPLOAD_DIR, f"export_{project.id}.xlsx")
    wb.save(filepath)
    return filepath


def export_csv(project, shots) -> str:
    filepath = os.path.join(settings.UPLOAD_DIR, f"export_{project.id}.csv")
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["镜头号", "镜头类型", "时长(秒)", "画面描述", "氛围", "对应脚本", "AI提示词", "备注"])
        for shot in shots:
            writer.writerow([
                shot.shot_number,
                shot.shot_type or "",
                shot.duration_sec or "",
                shot.content or "",
                shot.atmosphere or "",
                shot.script_reference or "",
                shot.ai_prompt or "",
                shot.notes or "",
            ])
    return filepath


def export_pdf(project, shots) -> BytesIO:
    """生成分镜表 PDF，返回 BytesIO buffer"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # 注册中文字体
    font_path = _find_cjk_font()
    if font_path:
        pdfmetrics.registerFont(TTFont("CJK", font_path))
    else:
        # 回退到 reportlab 内置 CID 字体
        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            font_path = "STSong-Light"
        except Exception:
            pass

    cjk_family = "CJK" if font_path and font_path != "STSong-Light" else "STSong-Light"

    buffer = BytesIO()

    # 根据分镜数量决定页面方向：>6 列用横向
    page_size = landscape(A4) if len(shots) > 6 else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        title=f"{project.title}_分镜表",
    )

    # 定义样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CNTitle",
        parent=styles["Title"],
        fontName=cjk_family,
        fontSize=14,
        leading=20,
        alignment=1,  # center
        spaceAfter=6 * mm,
    )
    cell_style = ParagraphStyle(
        "CNCell",
        parent=styles["Normal"],
        fontName=cjk_family,
        fontSize=7,
        leading=10,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "CNHeader",
        parent=cell_style,
        fontSize=8,
        leading=11,
        textColor=colors.white,
    )

    elements = []

    # 标题
    elements.append(Paragraph(f"{project.title} — 分镜表", title_style))

    # 表头
    headers = ["镜头号", "镜头类型", "时长(秒)", "画面描述", "氛围", "对应脚本", "AI提示词", "备注"]
    header_cells = [Paragraph(h, header_style) for h in headers]

    # 数据行
    data = [header_cells]
    for shot in shots:
        row = [
            Paragraph(str(shot.shot_number), cell_style),
            Paragraph(shot.shot_type or "", cell_style),
            Paragraph(str(shot.duration_sec) if shot.duration_sec else "", cell_style),
            Paragraph((shot.content or "").replace("\n", "<br/>"), cell_style),
            Paragraph(shot.atmosphere or "", cell_style),
            Paragraph((shot.script_reference or "").replace("\n", "<br/>"), cell_style),
            Paragraph((shot.ai_prompt or "").replace("\n", "<br/>"), cell_style),
            Paragraph(shot.notes or "", cell_style),
        ]
        data.append(row)

    # 计算列宽
    avail_width = page_size[0] - 20 * mm
    col_widths = [
        avail_width * 0.06,   # 镜头号
        avail_width * 0.07,   # 镜头类型
        avail_width * 0.06,   # 时长
        avail_width * 0.25,   # 画面描述
        avail_width * 0.10,   # 氛围
        avail_width * 0.20,   # 对应脚本
        avail_width * 0.18,   # AI提示词
        avail_width * 0.08,   # 备注
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # 表头样式
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        # 网格线
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, colors.HexColor("#2E5A9E")),
        # 对齐
        ("ALIGN", (0, 0), (2, -1), "CENTER"),    # 镜头号/类型/时长居中
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        # 斑马条纹
        *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F0F4FA" if i % 2 == 0 else "#FFFFFF"))
          for i in range(1, len(data))],
        # 内边距
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_images_zip(shots) -> str:
    filepath = os.path.join(settings.UPLOAD_DIR, f"images_{id(shots)}.zip")
    with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
        for shot in shots:
            if shot.reference_image_url:
                img_path = shot.reference_image_url.lstrip("/")
                full_path = os.path.join(settings.UPLOAD_DIR, "..", img_path)
                if os.path.exists(full_path):
                    zf.write(full_path, os.path.basename(full_path))
    return filepath
