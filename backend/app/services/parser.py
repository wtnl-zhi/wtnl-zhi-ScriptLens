import os


def parse_pdf(file_path: str) -> str:
    import fitz
    doc = fitz.open(file_path)
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text.strip()


def parse_docx(file_path: str) -> str:
    from docx import Document
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                texts.append(" ".join(cells))
    wb.close()
    return "\n".join(texts)


def parse_txt(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        return f.read().strip()


def parse_document(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    handlers = {
        ".pdf": parse_pdf,
        ".docx": parse_docx,
        ".xlsx": parse_xlsx,
        ".txt": parse_txt,
    }
    handler = handlers.get(ext)
    if not handler:
        raise ValueError(f"Unsupported file extension: {ext}")
    return handler(file_path)
